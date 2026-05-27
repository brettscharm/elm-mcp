"""Export DNG artifacts + attributes to a polished Excel workbook.

Used by the `export_module_to_xlsx` MCP tool. Pure helper — no I/O to ELM;
the caller passes requirement dicts already fetched via the client.

Output layout
-------------
- Summary sheet: project name, export timestamp, per-module counts.
- Either one sheet per module (default) or a single combined sheet with a
  Module column (when ``combined=True``).
- Header row is filled, bolded, frozen; auto-filter applied; columns
  auto-sized up to a max width.

Workbook files land in ``~/.elm-mcp/exports/`` by default.
"""
from __future__ import annotations
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_EXPORTS_DIR = Path.home() / ".elm-mcp" / "exports"

_BASE_COLUMNS: List[str] = ["ID", "Title", "Type", "URL"]
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _slug(s: str) -> str:
    cleaned = "".join(c if c.isalnum() or c in "-_" else "_" for c in s.strip())
    return cleaned[:60] or "export"


def _sheet_name(name: str, used: set) -> str:
    cleaned = "".join("_" if c in '/\\?*[]:' else c for c in name)[:31] or "Sheet"
    candidate = cleaned
    n = 2
    while candidate in used:
        suffix = f" ({n})"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


def _stringify(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (list, tuple)):
        return ", ".join(_stringify(x) for x in v)
    if isinstance(v, dict):
        return _stringify(v.get("label") or v.get("value") or v)
    return str(v)


def _attr_value(req: Dict[str, Any], key: str) -> str:
    if key in req and key not in {"custom_attributes"}:
        return _stringify(req[key])
    custom = req.get("custom_attributes") or {}
    if key in custom:
        return _stringify(custom[key])
    lower = key.lower()
    for k, v in custom.items():
        if k.lower() == lower:
            return _stringify(v)
    return ""


def export_artifacts_to_xlsx(
    modules: Sequence[Dict[str, Any]],
    project_name: str,
    columns: Optional[Sequence[str]] = None,
    combined: bool = False,
    output_dir: Optional[Path] = None,
) -> Path:
    """Write a polished xlsx for one DNG project.

    Args:
        modules: list of {"name": str, "requirements": [req dicts]}.
        project_name: DNG project title (filename + summary header).
        columns: attribute columns beyond ID/Title/Type/URL. If None,
            every attribute that appears in any requirement is included.
        combined: if True, all reqs go on one sheet with a Module column.
        output_dir: defaults to ``~/.elm-mcp/exports/``.

    Returns the path to the written file.
    """
    out_dir = output_dir or _EXPORTS_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    if columns is None:
        seen: List[str] = []
        for mod in modules:
            for req in mod.get("requirements", []):
                for k in (req.get("custom_attributes") or {}).keys():
                    if k not in seen and k not in _BASE_COLUMNS:
                        seen.append(k)
        columns = seen

    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Summary")
    summary.append(["Project", project_name])
    summary.append(["Exported", datetime.now().isoformat(timespec="seconds")])
    summary.append(["Modules", len(modules)])
    summary.append([
        "Total requirements",
        sum(len(m.get("requirements", [])) for m in modules),
    ])
    summary.append([])
    summary.append(["Module", "Requirements"])
    for mod in modules:
        summary.append([mod["name"], len(mod.get("requirements", []))])
    _format_header(summary, row=6)
    _autosize(summary)

    used_names: set = set()
    if combined:
        ws = wb.create_sheet(_sheet_name("All Requirements", used_names))
        ws.append(["Module"] + _BASE_COLUMNS + list(columns))
        for mod in modules:
            for req in mod.get("requirements", []):
                ws.append(_row_for(req, mod["name"], columns))
        _finalize(ws)
    else:
        for mod in modules:
            ws = wb.create_sheet(_sheet_name(mod["name"], used_names))
            ws.append(_BASE_COLUMNS + list(columns))
            for req in mod.get("requirements", []):
                ws.append(_row_for(req, None, columns))
            _finalize(ws)

    fname = f"{_slug(project_name)}-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    path = out_dir / fname
    wb.save(path)
    return path


def _row_for(
    req: Dict[str, Any],
    module_col: Optional[str],
    columns: Sequence[str],
) -> List[str]:
    row: List[str] = []
    if module_col is not None:
        row.append(module_col)
    row.extend([
        _stringify(req.get("id")),
        _stringify(req.get("title") or "(untitled)"),
        _stringify(req.get("artifact_type")),
        _stringify(req.get("url")),
    ])
    for col in columns:
        row.append(_attr_value(req, col))
    return row


def _format_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def _autosize(ws, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)


def _finalize(ws) -> None:
    _format_header(ws, row=1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)
